#!/usr/bin/env python3
"""
Batch Folder Resume Processing
===============================

Process all PDFs and DOCX files in a folder and export sections to Excel.
Only outputs filename and section content (no metadata).

Usage:
    python scripts/batch_folder_process.py --folder "freshteams_resume/Resumes" --output "outputs/batch_results.xlsx"
"""

import sys
from pathlib import Path

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

import argparse
import re
import time
from pathlib import Path
from typing import List, Dict, Any
from concurrent.futures import ThreadPoolExecutor, as_completed
from tqdm import tqdm

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("WARNING: openpyxl not installed. Install with: pip install openpyxl")

from src.core.unified_pipeline import UnifiedPipeline

# Filter for Excel-unsafe characters
ILLEGAL_CHARACTERS_RE = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]')


def sanitize_cell(value: str) -> str:
    """Sanitize cell value for Excel compatibility."""
    if value is None:
        return ""
    s = str(value)
    s = ILLEGAL_CHARACTERS_RE.sub("", s)
    return s.strip()


def find_resume_files(folder_path: str, recursive: bool = True) -> List[str]:
    """
    Find all PDF and DOCX files in a folder.
    
    Args:
        folder_path: Path to folder
        recursive: Search subfolders
        
    Returns:
        List of file paths
    """
    folder = Path(folder_path)
    
    if not folder.exists():
        print(f"Error: Folder not found: {folder_path}")
        return []
    
    files = []
    
    if recursive:
        # Search recursively
        files.extend(folder.rglob("*.pdf"))
        files.extend(folder.rglob("*.docx"))
        files.extend(folder.rglob("*.doc"))
    else:
        # Search only in the specified folder
        files.extend(folder.glob("*.pdf"))
        files.extend(folder.glob("*.docx"))
        files.extend(folder.glob("*.doc"))
    
    # Convert to strings and sort
    file_paths = sorted([str(f) for f in files])
    
    return file_paths


def process_single_file(file_path: str, pipeline: UnifiedPipeline) -> Dict[str, Any]:
    """Process a single resume file."""
    try:
        result = pipeline.parse(file_path, verbose=False)
        return result
    except Exception as e:
        return {
            'success': False,
            'metadata': {
                'file_path': file_path,
                'file_name': Path(file_path).name,
                'error': str(e)
            }
        }


def save_to_excel(results: List[Dict[str, Any]], output_path: str):
    """
    Save results to Excel with just filename and section content.
    
    Args:
        results: List of processing results
        output_path: Path to Excel file
    """
    if not OPENPYXL_AVAILABLE:
        print("Error: openpyxl not installed. Cannot save to Excel.")
        print("Install with: pip install openpyxl")
        return
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Resume Sections"
    
    # Define standard sections in preferred order
    standard_sections = [
        "Contact Information",
        "Summary",
        "Skills",
        "Experience",
        "Projects",
        "Education",
        "Certifications",
        "Achievements",
        "Publications",
        "Research",
        "Languages",
        "Volunteer",
        "Hobbies",
        "References",
        "Declarations",
        "Unknown Sections"
    ]
      # Header row - pdf_path (for web viewer), file name, and sections
    headers = ["pdf_path", "File Name"] + standard_sections
    ws.append(headers)
    
    # Style header
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
      # Data rows
    for result in results:
        if not result.get('success'):
            # Failed processing - just show error
            pdf_path = sanitize_cell(result['metadata'].get('file_path', ''))
            file_name = result['metadata'].get('file_name', 'Unknown')
            error_msg = result['metadata'].get('error', 'Unknown error')
            row = [sanitize_cell(pdf_path), sanitize_cell(file_name)] + [sanitize_cell(f"ERROR: {error_msg}")] + [""] * (len(standard_sections) - 1)
            ws.append(row)
            continue
        
        metadata = result.get('metadata', {})
        parsed_data = result.get('result', {})
        sections_data = parsed_data.get('sections', [])
          # Get pdf_path (full path for web viewer) and file name
        pdf_path = sanitize_cell(metadata.get('file_path', ''))
        file_name = sanitize_cell(metadata.get('file_name', ''))
          # Create a map of section name to content
        section_map = {}
        for section in sections_data:
            section_name = section.get('section', '')
            lines = section.get('lines', [])
            
            # Clean up lines - remove any dict/metadata formatting
            clean_lines = []
            for line in lines:
                if not line:
                    continue
                
                # Extract text from line object if it's a dict
                if isinstance(line, dict):
                    line_str = str(line.get('text', '')).strip()
                else:
                    line_str = str(line).strip()
                
                if not line_str:
                    continue
                
                # Skip lines that look like metadata (dict format, JSON, etc.)
                if line_str.startswith('{') and line_str.endswith('}'):
                    # Try to parse as JSON and extract values
                    try:
                        import json
                        data = json.loads(line_str)
                        if isinstance(data, dict):
                            # Extract only the values, not the keys
                            for key, value in data.items():
                                if value and str(value).strip():
                                    clean_lines.append(str(value).strip())
                            continue
                    except:
                        pass
                
                # Skip lines with metadata markers
                if any(marker in line_str for marker in ['◆', '❖', '✦', '⚬']):
                    # Remove the markers
                    line_str = line_str.replace('◆', '').replace('❖', '').replace('✦', '').replace('⚬', '').strip()
                
                if line_str:
                    clean_lines.append(line_str)
            
            # Join lines with newline for better readability in Excel
            content = '\n'.join(clean_lines)
            section_map[section_name] = content
        
        # Build row: pdf_path + file name + section contents
        row = [pdf_path, file_name]
        for section_name in standard_sections:
            content = section_map.get(section_name, '')
            row.append(sanitize_cell(content))
        
        ws.append(row)
        
        # Enable text wrapping for content cells
        for cell in ws[ws.max_row]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
      # Set column widths
    ws.column_dimensions['A'].width = 50  # pdf_path column (hidden but needed for web viewer)
    ws.column_dimensions['B'].width = 30  # File name column
    for col_idx in range(3, len(headers) + 1):
        col_letter = ws.cell(1, col_idx).column_letter
        ws.column_dimensions[col_letter].width = 45  # Section columns
    
    # Hide pdf_path column (keep it for web viewer but hide from users)
    ws.column_dimensions['A'].hidden = True
    
    # Freeze first row and second column (File Name)
    ws.freeze_panes = 'C2'
    
    # Save workbook
    wb.save(output_path)
    print(f"\n✅ Excel saved: {output_path}")
    print(f"   Total rows: {len(results)}")
    print(f"   Columns: {len(headers)}")


def process_folder(
    folder_path: str,
    output_excel: str,
    max_workers: int = 4,
    recursive: bool = True,
    verbose: bool = True
):
    """
    Process all resumes in a folder and save to Excel.
    
    Args:
        folder_path: Path to folder containing resumes
        output_excel: Path to output Excel file
        max_workers: Number of parallel workers
        recursive: Search subfolders
        verbose: Print progress
    """
    start_time = time.time()
    
    # Find all resume files
    if verbose:
        print(f"\n🔍 Searching for resumes in: {folder_path}")
        print(f"   Recursive: {recursive}")
    
    file_paths = find_resume_files(folder_path, recursive=recursive)
    
    if not file_paths:
        print(f"\n❌ No resume files found in: {folder_path}")
        return
    
    if verbose:
        print(f"\n📄 Found {len(file_paths)} files")
        print(f"   Workers: {max_workers}")
      # Initialize pipeline with verbose=False to suppress section output
    pipeline = UnifiedPipeline(enable_learning=True, verbose=False)
    
    # Process files in parallel
    results = []
    
    if verbose:
        print(f"\n⚙️  Processing resumes...")
        print(f"   Note: Section details hidden for clean output. Check Excel for results.")
    
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        # Submit all tasks
        future_to_file = {
            executor.submit(process_single_file, path, pipeline): path
            for path in file_paths
        }
        
        # Process with progress bar
        with tqdm(total=len(file_paths), desc="Progress", disable=not verbose) as pbar:
            for future in as_completed(future_to_file):
                try:
                    result = future.result()
                    results.append(result)
                except Exception as e:
                    file_path = future_to_file[future]
                    results.append({
                        'success': False,
                        'metadata': {
                            'file_path': file_path,
                            'file_name': Path(file_path).name,
                            'error': str(e)
                        }
                    })
                
                pbar.update(1)
    
    # Calculate statistics
    successful = sum(1 for r in results if r.get('success'))
    failed = len(results) - successful
    total_time = time.time() - start_time
    
    if verbose:
        print(f"\n📊 Processing Complete:")
        print(f"   Total: {len(results)}")
        print(f"   Success: {successful} ({successful/len(results)*100:.1f}%)")
        print(f"   Failed: {failed} ({failed/len(results)*100:.1f}%)")
        print(f"   Time: {total_time:.2f}s ({total_time/len(results):.2f}s/file)")
    
    # Save to Excel
    if verbose:
        print(f"\n💾 Saving to Excel...")
    
    save_to_excel(results, output_excel)
    
    if verbose:
        print(f"\n✨ Done!")


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description="Process all resumes in a folder and export sections to Excel"
    )
    parser.add_argument(
        "--folder",
        required=True,
        help="Path to folder containing resumes"
    )
    parser.add_argument(
        "--output",
        default="outputs/batch_results.xlsx",
        help="Path to output Excel file (default: outputs/batch_results.xlsx)"
    )
    parser.add_argument(
        "--workers",
        type=int,
        default=4,
        help="Number of parallel workers (default: 4)"
    )
    parser.add_argument(
        "--no-recursive",
        action="store_true",
        help="Don't search subfolders"
    )
    parser.add_argument(
        "--quiet",
        action="store_true",
        help="Quiet mode (no progress output)"
    )
    
    args = parser.parse_args()
    
    # Create output directory if it doesn't exist
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Process folder
    process_folder(
        folder_path=args.folder,
        output_excel=args.output,
        max_workers=args.workers,
        recursive=not args.no_recursive,
        verbose=not args.quiet
    )


if __name__ == "__main__":
    main()
