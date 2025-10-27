#!/usr/bin/env python3
"""
Web-based Resume Parser Results Viewer
=======================================

View parsed resume results in a browser with side-by-side comparison.

Usage:
    python scripts/view_results.py
    # Then open browser to http://localhost:5000
"""

import sys
from pathlib import Path

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

import os
import json
from pathlib import Path
from typing import List, Dict, Any

from flask import Flask, jsonify, send_file, abort, Response, request
from openpyxl import load_workbook

from src.PDF_pipeline.segment_sections import SECTIONS

# Config
# Default Excel file name as requested. Fallback to outputs/batch_sections.xlsx if missing.
DEFAULT_XLSX = "outputs/batch_results.xlsx"
FALLBACK_XLSX = "outputs/batch_results.xlsx"
ALSO_TRY_XLSX = "outputs/batch_results.xlsx"  # batch writer default name

# Resolve project root (this file is in scripts/ folder, so go up one level)
ROOT_DIR = Path(__file__).resolve().parent.parent
INDEX_HTML = Path(__file__).resolve().parent / "index.html"

def _windows_to_wsl_path(s: str) -> Path:
    """
    Convert Windows absolute path like C:\\Users\\... to WSL /mnt/c/Users/...
    If not a drive path, return as Path(s).
    """
    if not s:
        return Path("")
    # Drive letter pattern: X:\ or X:/
    if len(s) >= 3 and s[1] == ":" and (s[2] == "\\" or s[2] == "/") and s[0].isalpha():
        drive = s[0].lower()
        rest = s[2:].replace("\\", "/").lstrip("/")
        return Path(f"/mnt/{drive}/{rest}")
    return Path(s)

def _sanitize_path_str(raw: str) -> str:
    # Remove common prefixes like file://
    if not raw:
        return raw
    if raw.startswith("file://"):
        return raw.replace("file://", "", 1)
    return raw

def _search_workspace_for(name: str) -> Path | None:
    """Search the workspace for a file by exact filename, then by stem (preferring .pdf)."""
    try:
        # Exact filename match
        matches = list(ROOT_DIR.rglob(name))
        if matches:
            # Prefer PDFs then shorter paths
            matches.sort(key=lambda p: (p.suffix.lower() != ".pdf", len(str(p))))
            return matches[0]
        # By stem
        stem = Path(name).stem
        if not stem:
            return None
        stem_matches = [p for p in ROOT_DIR.rglob(f"{stem}.*")]
        if stem_matches:
            stem_matches.sort(key=lambda p: (p.suffix.lower() != ".pdf", len(str(p))))
            return stem_matches[0]
    except Exception:
        pass
    return None

def _load_rows_from_excel(xlsx_path: Path) -> List[Dict[str, Any]]:
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Excel file not found: {xlsx_path}")

    wb = load_workbook(str(xlsx_path), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(h or "").strip() for h in rows[0]]
    data_rows = rows[1:]
    cols = {name: idx for idx, name in enumerate(header)}

    # Required columns
    if "File Name" not in cols:
        # Fallback to old format with pdf_path
        if "pdf_path" not in cols:
            raise ValueError(f"Missing required column 'File Name' or 'pdf_path' in Excel header: {header}")
        file_path_col = "pdf_path"
    else:
        file_path_col = "File Path" if "File Path" in cols else "File Name"

    # Identify section columns - anything that's not File Name or File Path
    section_cols_in_sheet = [h for h in header if h not in ["File Name", "File Path", "pdf_path", "Error"]]
    
    items: List[Dict[str, Any]] = []
    for i, r in enumerate(data_rows):
        if r is None:
            continue
        
        # Get file path
        file_name_idx = cols.get("File Name", cols.get("pdf_path", 0))
        file_path_idx = cols.get(file_path_col, file_name_idx)
        
        file_name = (r[file_name_idx] or "").strip() if len(r) > file_name_idx else ""
        file_path = (r[file_path_idx] or "").strip() if len(r) > file_path_idx else ""
        
        if not file_name and not file_path:
            continue
        
        # Use file_path if available, otherwise use file_name
        pdf_path = file_path if file_path else file_name
        
        # Parse contact info from "Contact Information" section if available
        contact: Dict[str, Any] = {}
        contact_info_idx = cols.get("Contact Information")
        if contact_info_idx is not None and len(r) > contact_info_idx:
            contact_raw = r[contact_info_idx] or ""
            if isinstance(contact_raw, str) and contact_raw.strip():
                # Parse contact lines to extract structured info
                lines = contact_raw.split('\n')
                for line in lines:
                    line = line.strip()
                    if '@' in line:
                        contact['email'] = line
                    elif any(c.isdigit() for c in line) and len(line) > 8:
                        if 'phone' not in contact:
                            contact['phone'] = line
                    elif 'linkedin.com' in line.lower():
                        contact['linkedin'] = line
                    elif 'github.com' in line.lower():
                        contact['github'] = line
                    elif not contact.get('name') and len(line.split()) <= 4:
                        # First short line might be name
                        contact['name'] = line

        sections: Dict[str, Any] = {}
        for sec in section_cols_in_sheet:
            idx = cols.get(sec)
            if idx is None or idx >= len(r):
                continue
            val = r[idx] or ""
            # Section lines were joined with '\n' in batch writer; split to list for UI
            if isinstance(val, str):
                lines = [ln.strip() for ln in val.split("\n") if str(ln).strip()]
            elif isinstance(val, (list, tuple)):
                lines = [str(v).strip() for v in val if str(v).strip()]
            else:
                lines = [str(val).strip()] if str(val).strip() else []
            if lines:
                sections[sec] = lines

        items.append({
            "index": i,
            "pdf_path": pdf_path,
            "contact": contact,
            "sections": sections,
        })
    return items

def _load_rows_from_json(json_path: Path) -> List[Dict[str, Any]]:
    """Load resume data from JSON format (batch_results_*.json)"""
    if not json_path.exists():
        raise FileNotFoundError(f"JSON file not found: {json_path}")
    
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    results = data.get("results", [])
    items: List[Dict[str, Any]] = []
    
    for i, entry in enumerate(results):
        if not entry.get("success"):
            continue
        
        result = entry.get("result", {})
        file_path = entry.get("file_path", "")
        
        if not file_path:
            continue
        
        # Extract contact information (if available)
        contact: Dict[str, Any] = {}
        
        # Parse sections
        sections_data = result.get("sections", [])
        sections: Dict[str, Any] = {}
        
        for section in sections_data:
            section_name = section.get("section_name", "Unknown")
            content = section.get("content", "")
            
            # Special handling for contact information in Education or other sections
            if section_name == "Contact Information" or "contact" in section_name.lower():
                if isinstance(content, str) and content.strip():
                    lines = content.split('\n')
                    for line in lines:
                        line = line.strip()
                        if '@' in line:
                            contact['email'] = line
                        elif any(c.isdigit() for c in line) and len(line) > 8:
                            if 'phone' not in contact:
                                contact['phone'] = line
                        elif 'linkedin.com' in line.lower():
                            contact['linkedin'] = line
                        elif 'github.com' in line.lower():
                            contact['github'] = line
            
            # Convert content to list of lines
            if isinstance(content, str):
                lines = [ln.strip() for ln in content.split("\n") if ln.strip()]
            elif isinstance(content, list):
                lines = [str(v).strip() for v in content if str(v).strip()]
            else:
                lines = [str(content).strip()] if str(content).strip() else []
            
            if lines:
                sections[section_name] = lines
        
        items.append({
            "index": i,
            "pdf_path": file_path,
            "contact": contact,
            "sections": sections,
        })
    
    return items

def _resolve_excel_path() -> Path:
    # Allow override via env
    env_val = os.getenv("BATCH_XLSX")
    if env_val:
        p = Path(env_val)
        if not p.is_absolute():
            p = ROOT_DIR / p
        if p.exists():
            return p

    candidates = [
        ROOT_DIR / DEFAULT_XLSX,
        ROOT_DIR / ALSO_TRY_XLSX,
        ROOT_DIR / FALLBACK_XLSX,
        ROOT_DIR / "outputs" / DEFAULT_XLSX,
    ]
    for p in candidates:
        if p.exists():
            return p

    # last resort: look in root for any xlsx containing 'batch' in name
    for cand in ROOT_DIR.glob("*.xlsx"):
        if "batch" in cand.name.lower():
            return cand
    raise FileNotFoundError(f"Could not find {DEFAULT_XLSX} or {ALSO_TRY_XLSX} or {FALLBACK_XLSX} in {ROOT_DIR}")

def _resolve_data_path() -> tuple[Path, str]:
    """Returns (path, type) where type is 'json' or 'excel'"""
    # Allow override via env
    env_val = os.getenv("BATCH_DATA")
    if env_val:
        p = Path(env_val)
        if not p.is_absolute():
            p = ROOT_DIR / p
        if p.exists():
            return (p, "json" if p.suffix == ".json" else "excel")

    # Try JSON files first (newer format)
    json_candidates = [
        ROOT_DIR / "outputs" / "batch_results_20251027_163813.json",
        ROOT_DIR / "batch_results.json",
        ROOT_DIR / "outputs" / "batch_results.json",
    ]
    for p in json_candidates:
        if p.exists():
            return (p, "json")
    
    # Look for any JSON in outputs with batch_results in name
    outputs_dir = ROOT_DIR / "outputs"
    if outputs_dir.exists():
        for cand in outputs_dir.glob("batch_results*.json"):
            return (cand, "json")

    # Fall back to Excel
    excel_candidates = [
        ROOT_DIR / DEFAULT_XLSX,
        ROOT_DIR / ALSO_TRY_XLSX,
        ROOT_DIR / FALLBACK_XLSX,
        ROOT_DIR / "outputs" / DEFAULT_XLSX,
    ]
    for p in excel_candidates:
        if p.exists():
            return (p, "excel")

    # Last resort: look in root for any xlsx containing 'batch' in name
    for cand in ROOT_DIR.glob("*.xlsx"):
        if "batch" in cand.name.lower():
            return (cand, "excel")
    
    raise FileNotFoundError(f"Could not find batch_results JSON or Excel file in {ROOT_DIR}")

def _build_allowed_pdf_map(items: List[Dict[str, Any]]) -> Dict[int, Path]:
    """
    Map index -> absolute Path. Only serve files that appear in the Excel.
    Also translate Windows-style absolute paths to WSL if needed. Fall back to workspace search.
    """
    idx_to_path: Dict[int, Path] = {}
    for i, item in enumerate(items):
        raw = str(item.get("pdf_path", "") or "").strip()
        if not raw:
            continue
        raw = _sanitize_path_str(raw)
        p = Path(raw)

        # Prefer absolute resolution or relative to project
        if not p.is_absolute():
            p = (ROOT_DIR / raw).resolve()

        # If missing, try Windows->WSL translation
        if not p.exists():
            p2 = _windows_to_wsl_path(raw)
            if p2.exists():
                p = p2

        # If still missing, search workspace by filename/stem, prefer PDF
        if not p.exists():
            fallback = _search_workspace_for(Path(raw).name)
            if fallback and fallback.exists():
                p = fallback

        idx_to_path[i] = p

        # Optional: log missing for debugging
        if not p.exists():
            try:
                print(f"[warn] Missing file for row {i}: '{raw}' -> resolved '{p}'", flush=True)
            except Exception:
                pass
    return idx_to_path

def create_app() -> Flask:
    app = Flask(__name__)

    # Load data once at startup - support both JSON and Excel formats
    data_path, data_type = _resolve_data_path()
    
    if data_type == "json":
        items = _load_rows_from_json(data_path)
        print(f"[info] Loaded {len(items)} items from JSON: {data_path}", flush=True)
    else:
        items = _load_rows_from_excel(data_path)
        print(f"[info] Loaded {len(items)} items from Excel: {data_path}", flush=True)
    
    path_map = _build_allowed_pdf_map(items)

    @app.get("/")
    def index() -> Response:
        if not INDEX_HTML.exists():
            return Response("index.html not found at project root.", status=404)
        return send_file(str(INDEX_HTML), mimetype="text/html; charset=utf-8")

    @app.get("/api/rows")
    def api_rows():
        # Filtering controlled via query params. Defaults to no filtering (all).
        # New param: experience={all|empty|nonempty}. Back-compat: experience_empty={0|1}
        q_mode = request.args.get("experience")
        q_legacy = request.args.get("experience_empty")

        def has_empty_experience(item: Dict[str, Any]) -> bool:
            secs = item.get("sections", {}) or {}
            # find the experience key case-insensitively
            exp_key = None
            for k in secs.keys():
                if str(k).strip().lower() == "experience":
                    exp_key = k
                    break
            exp_lines = secs.get(exp_key) if exp_key is not None else None
            # treat missing or empty list/string as empty
            if exp_lines is None:
                return True
            if isinstance(exp_lines, str):
                return len(exp_lines.strip()) == 0
            if isinstance(exp_lines, list):
                return len([ln for ln in exp_lines if str(ln).strip()]) == 0
            return False

        # Determine filter mode
        mode = "all"
        if q_mode:
            v = q_mode.strip().lower()
            if v in ("empty", "none"):
                mode = "empty"
            elif v in ("nonempty", "notnull", "not-null"):
                mode = "nonempty"
            else:
                mode = "all"
        elif q_legacy is not None:
            # legacy boolean flag
            mode = "empty" if q_legacy.strip().lower() in ("1", "true", "yes") else "all"

        payload = []
        for i, item in enumerate(items):
            empty = has_empty_experience(item)
            if mode == "empty" and not empty:
                continue
            if mode == "nonempty" and empty:
                continue
            payload.append({
                "i": i,
                "pdf_path": item["pdf_path"],
                "filename": Path(item["pdf_path"]).name,                "contact": item.get("contact", {}),
                "sections": item.get("sections", {}),
            })
        return jsonify(payload)

    @app.get("/pdf/<int:i>")
    def serve_pdf_by_index(i: int):
        if i < 0 or i >= len(items):
            abort(404)
        p = path_map.get(i)
        if not p or not p.exists():
            abort(404)
        # Determine mimetype
        ext = p.suffix.lower()
        if ext == ".pdf":
            mimetype = "application/pdf"
        elif ext in (".png", ".jpg", ".jpeg"):
            mimetype = f"image/{'jpeg' if ext in ('.jpg', '.jpeg') else 'png'}"
        elif ext in (".doc", ".docx"):
            # For Word docs, send as download with proper mimetype
            mimetype = "application/vnd.openxmlformats-officedocument.wordprocessingml.document" if ext == ".docx" else "application/msword"
        else:
            # default binary; iframe may not render but still allows download
            mimetype = "application/octet-stream"
        return send_file(str(p), mimetype=mimetype, conditional=True)

    return app

if __name__ == "__main__":
    app = create_app()
    app.run(host='0.0.0.0', port=5000, debug=True)

