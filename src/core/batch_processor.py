"""
Batch Resume Processing
=======================

Process multiple resumes in parallel with progress tracking,
error handling, and automatic result aggregation.
"""

import time
import json
import re
from pathlib import Path
from typing import List, Dict, Any, Optional
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from tqdm import tqdm

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from src.core.unified_pipeline import UnifiedPipeline

# Filter for Excel-unsafe characters
ILLEGAL_CHARACTERS_RE = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]')


@dataclass
class BatchResult:
    """Result of batch processing."""
    total: int
    successful: int
    failed: int
    results: List[Dict[str, Any]]
    errors: List[Dict[str, str]]
    processing_time: float
    avg_time_per_file: float


class BatchProcessor:
    """
    Process multiple resumes in parallel.
    """
    
    def __init__(
        self,
        config_path: Optional[str] = None,
        max_workers: int = 4,
        enable_learning: bool = True,
        verbose: bool = False
    ):
        """
        Initialize batch processor.
        
        Args:
            config_path: Path to sections config
            max_workers: Number of parallel workers
            enable_learning: Enable section learning
            verbose: Print detailed progress
        """
        self.max_workers = max_workers
        self.verbose = verbose
        self.enable_learning = enable_learning
        
        # Initialize pipeline
        self.pipeline = UnifiedPipeline(
            config_path=config_path,
            enable_learning=enable_learning,
            verbose=False  # Disable verbose for parallel processing
        )
        
        if self.verbose:
            print(f"[Batch Processor] Initialized with {max_workers} workers")
    
    def process_directory(
        self,
        directory: str,
        pattern: str = "*.pdf",
        output_dir: Optional[str] = None,
        save_individual: bool = False
    ) -> BatchResult:
        """
        Process all resumes in a directory.
        
        Args:
            directory: Directory containing resumes
            pattern: File pattern to match (e.g., "*.pdf", "*.docx")
            output_dir: Directory to save results
            save_individual: Save individual JSON files
            
        Returns:
            BatchResult object with statistics and results
        """
        start_time = time.time()
        
        # Find all matching files
        files = list(Path(directory).glob(pattern))
        
        if not files:
            if self.verbose:
                print(f"No files matching '{pattern}' found in {directory}")
            
            return BatchResult(
                total=0,
                successful=0,
                failed=0,
                results=[],
                errors=[],
                processing_time=0.0,
                avg_time_per_file=0.0
            )
        
        if self.verbose:
            print(f"\n{'='*60}")
            print(f"[Batch Processing] {len(files)} files")
            print(f"{'='*60}\n")
        
        # Process files in parallel
        results = []
        errors = []
        
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            # Submit all tasks
            future_to_file = {
                executor.submit(self._process_single, str(f)): f
                for f in files
            }
            
            # Process completed tasks with progress bar
            with tqdm(total=len(files), desc="Processing", disable=not self.verbose) as pbar:
                for future in as_completed(future_to_file):
                    file_path = future_to_file[future]
                    
                    try:
                        result = future.result()
                        
                        if result['success']:
                            results.append(result)
                        else:
                            errors.append({
                                'file': str(file_path),
                                'error': result['metadata'].get('error', 'Unknown error')
                            })
                    
                    except Exception as e:
                        errors.append({
                            'file': str(file_path),
                            'error': str(e)
                        })
                    
                    pbar.update(1)
        
        # Save results if output directory specified
        if output_dir and results:
            output_path = Path(output_dir)
            output_path.mkdir(parents=True, exist_ok=True)
            
            # Save aggregated results
            summary_path = output_path / "batch_summary.json"
            self._save_summary(results, errors, summary_path)
            
            # Save individual files if requested
            if save_individual:
                for result in results:
                    file_name = Path(result['metadata']['file_name']).stem
                    individual_path = output_path / f"{file_name}.json"
                    
                    with open(individual_path, 'w', encoding='utf-8') as f:
                        json.dump(result['result'], f, ensure_ascii=False, indent=2)
            
            if self.verbose:
                print(f"\n[Saved] Results to: {output_dir}")
        
        # Calculate statistics
        processing_time = time.time() - start_time
        successful = len(results)
        failed = len(errors)
        total = len(files)
        avg_time = processing_time / total if total > 0 else 0.0
        
        if self.verbose:
            print(f"\n{'='*60}")
            print(f"[Batch Complete]")
            print(f"  Total: {total}")
            print(f"  Success: {successful} ({successful/total*100:.1f}%)")
            print(f"  Failed: {failed} ({failed/total*100:.1f}%)")
            print(f"  Time: {processing_time:.2f}s")
            print(f"  Avg: {avg_time:.2f}s/file")
            print(f"{'='*60}\n")
        
        return BatchResult(
            total=total,
            successful=successful,
            failed=failed,
            results=results,
            errors=errors,
            processing_time=processing_time,
            avg_time_per_file=avg_time
        )
    
    def process_files(
        self,
        file_paths: List[str],
        output_dir: Optional[str] = None,
        save_individual: bool = False
    ) -> BatchResult:
        """
        Process a list of specific files.
        
        Args:
            file_paths: List of file paths to process
            output_dir: Directory to save results
            save_individual: Save individual JSON files
            
        Returns:
            BatchResult object
        """
        start_time = time.time()
        
        if self.verbose:
            print(f"\n{'='*60}")
            print(f"[Batch Processing] {len(file_paths)} files")
            print(f"{'='*60}\n")
        
        results = []
        errors = []
        
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            future_to_file = {
                executor.submit(self._process_single, f): f
                for f in file_paths
            }
            
            with tqdm(total=len(file_paths), desc="Processing", disable=not self.verbose) as pbar:
                for future in as_completed(future_to_file):
                    file_path = future_to_file[future]
                    
                    try:
                        result = future.result()
                        
                        if result['success']:
                            results.append(result)
                        else:
                            errors.append({
                                'file': file_path,
                                'error': result['metadata'].get('error', 'Unknown error')
                            })
                    
                    except Exception as e:
                        errors.append({
                            'file': file_path,
                            'error': str(e)
                        })
                    
                    pbar.update(1)
        
        # Save results if requested
        if output_dir and results:
            output_path = Path(output_dir)
            output_path.mkdir(parents=True, exist_ok=True)
            
            summary_path = output_path / "batch_summary.json"
            self._save_summary(results, errors, summary_path)
            
            if save_individual:
                for result in results:
                    file_name = Path(result['metadata']['file_name']).stem
                    individual_path = output_path / f"{file_name}.json"
                    
                    with open(individual_path, 'w', encoding='utf-8') as f:
                        json.dump(result['result'], f, ensure_ascii=False, indent=2)
        
        processing_time = time.time() - start_time
        successful = len(results)
        failed = len(errors)
        total = len(file_paths)
        avg_time = processing_time / total if total > 0 else 0.0
        
        if self.verbose:
            print(f"\n{'='*60}")
            print(f"[Batch Complete]")
            print(f"  Success: {successful}/{total}")
            print(f"  Failed: {failed}/{total}")
            print(f"  Time: {processing_time:.2f}s")
            print(f"  Avg: {avg_time:.2f}s/file")
            print(f"{'='*60}\n")
        
        return BatchResult(
            total=total,
            successful=successful,
            failed=failed,
            results=results,
            errors=errors,
            processing_time=processing_time,
            avg_time_per_file=avg_time
        )
    
    def process_batch(
        self,
        input_paths: List[str],
        output_excel: Optional[str] = None,
        max_workers: Optional[int] = None,
        verbose: bool = False
    ) -> Dict[str, Any]:
        """
        Process multiple resumes in parallel.
        
        Args:
            input_paths: List of file paths to process
            output_excel: Optional Excel output file
            max_workers: Override max_workers setting
            verbose: Print progress
            
        Returns:
            Dictionary with:
                - total_processed: int
                - successful: int
                - failed: int
                - results: List[Dict]
                - total_time: float
        """
        start_time = time.time()
        
        if max_workers:
            workers = max_workers
        else:
            workers = self.max_workers
        
        if verbose:
            print(f"\nProcessing {len(input_paths)} files with {workers} workers...")
        
        results = []
        errors = []
        
        with ThreadPoolExecutor(max_workers=workers) as executor:
            future_to_file = {
                executor.submit(self._process_single, path): path
                for path in input_paths
            }
            
            for future in as_completed(future_to_file):
                file_path = future_to_file[future]
                
                try:
                    result = future.result()
                    
                    if result['success']:
                        results.append(result)
                    else:
                        errors.append({
                            'file': file_path,
                            'error': result['metadata'].get('error', 'Unknown')
                        })
                
                except Exception as e:
                    errors.append({
                        'file': file_path,
                        'error': str(e)
                    })
        
        # Save to Excel if requested
        if output_excel and results:
            self._save_to_excel(results, output_excel)
        
        total_time = time.time() - start_time
        
        return {
            'total_processed': len(input_paths),
            'successful': len(results),
            'failed': len(errors),
            'results': results,
            'errors': errors,
            'total_time': total_time,
            'avg_time': total_time / len(input_paths) if input_paths else 0
        }
    
    @staticmethod
    def _sanitize_cell(value: str) -> str:
        """Sanitize cell value for Excel compatibility."""
        if value is None:
            return ""
        s = str(value)
        s = ILLEGAL_CHARACTERS_RE.sub("", s)
        return s.strip()
    
    def _save_to_excel(
        self,
        results: List[Dict[str, Any]],
        output_path: str,
        include_sections: bool = True
    ):
        """
        Save results to Excel file with sections as columns.
        
        Args:
            results: List of processing results
            output_path: Path to Excel file
            include_sections: Include individual sections as columns
        """
        if not OPENPYXL_AVAILABLE:
            print("Warning: openpyxl not available. Install with: pip install openpyxl")
            return
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Resume Data"
        
        # Define all potential section names from results
        all_sections = set()
        for result in results:
            if result.get('success') and result.get('result'):
                sections = result['result'].get('sections', [])
                for section in sections:
                    section_name = section.get('section', 'Unknown')
                    all_sections.add(section_name)
        
        # Sort sections for consistent column order
        section_columns = sorted(all_sections)
        
        # Build header row
        headers = ['File Name', 'Success', 'Strategy', 'Processing Time (s)', 'Total Sections', 'Contact Info']
        if include_sections:
            headers.extend(section_columns)
        
        # Style header
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add data rows
        for row_idx, result in enumerate(results, 2):
            file_name = Path(result['metadata']['file_name']).name
            success = result.get('success', False)
            strategy = result.get('strategy', 'unknown')
            proc_time = result['metadata'].get('processing_time', 0)
            
            # Basic columns
            ws.cell(row=row_idx, column=1, value=self._sanitize_cell(file_name))
            ws.cell(row=row_idx, column=2, value='✓' if success else '✗')
            ws.cell(row=row_idx, column=3, value=self._sanitize_cell(strategy))
            ws.cell(row=row_idx, column=4, value=round(proc_time, 2))
            
            if success and result.get('result'):
                data = result['result']
                sections = data.get('sections', [])
                
                # Total sections count
                ws.cell(row=row_idx, column=5, value=len(sections))
                
                # Contact info (as JSON string)
                contact = data.get('contact', {})
                contact_str = json.dumps(contact, ensure_ascii=False) if contact else ''
                ws.cell(row=row_idx, column=6, value=self._sanitize_cell(contact_str))
                
                if include_sections:
                    # Create a map of section name to content
                    section_map = {}
                    for section in sections:
                        section_name = section.get('section', 'Unknown')
                        lines = section.get('lines', [])
                        # Join lines into single text
                        content = '\n'.join(str(line) for line in lines if line)
                        section_map[section_name] = content
                    
                    # Fill section columns
                    for col_idx, section_name in enumerate(section_columns, 7):
                        content = section_map.get(section_name, '')
                        ws.cell(row=row_idx, column=col_idx, value=self._sanitize_cell(content))
            else:
                # Failed processing
                ws.cell(row=row_idx, column=5, value=0)
                error_msg = result['metadata'].get('error', 'Unknown error')
                ws.cell(row=row_idx, column=6, value=self._sanitize_cell(f"Error: {error_msg}"))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # Set width (max 50 for readability)
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze first row
        ws.freeze_panes = 'A2'
        
        # Save workbook
        wb.save(output_path)
        
        if self.verbose:
            print(f"  Excel file saved: {output_path}")

    def _process_single(self, file_path: str) -> Dict[str, Any]:
        """Process a single file."""
        return self.pipeline.parse(file_path)
    
    def _save_summary(
        self,
        results: List[Dict[str, Any]],
        errors: List[Dict[str, str]],
        output_path: Path
    ):
        """Save batch processing summary."""
        summary = {
            'total_files': len(results) + len(errors),
            'successful': len(results),
            'failed': len(errors),
            'timestamp': time.strftime('%Y-%m-%d %H:%M:%S'),
            'results': [
                {
                    'file': r['metadata']['file_name'],
                    'sections': len(r['result'].get('sections', [])),
                    'lines': sum(len(s.get('lines', [])) for s in r['result'].get('sections', [])),
                    'strategy': r['strategy'],
                    'processing_time': r['metadata']['processing_time']
                }
                for r in results
            ],
            'errors': errors
        }
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(summary, f, ensure_ascii=False, indent=2)
