import os
import json
from pathlib import Path
from typing import List, Dict, Any

from flask import Flask, jsonify, send_file, abort, Response
from openpyxl import load_workbook

from src.PDF_pipeline.segment_sections import SECTIONS

# Config
# Default Excel file name as requested. Fallback to outputs/batch_sections.xlsx if missing.
DEFAULT_XLSX = "batch_selection_4.xlsx"
FALLBACK_XLSX = "outputs/batch_sections_4.xlsx"
ALSO_TRY_XLSX = "batch_sections_4.xlsx"  # batch writer default name

# Resolve project root (this file is in project root)
ROOT_DIR = Path(__file__).resolve().parent
INDEX_HTML = ROOT_DIR / "index.html"

def _windows_to_wsl_path(s: str) -> Path:
    """
    Convert Windows absolute path like C:\\Users\\... to WSL /mnt/c/Users/...
    If not a drive path, return as Path(s).
    """
    if not s:
        return Path("")
    # Drive letter pattern: X:\ or X:/
    if len(s) >= 3 and s[1] == ":" and (s[2] == "\\" or s[2] == "/") and s[0].isalpha():
        drive = s[0].lower()
        rest = s[2:].replace("\\", "/").lstrip("/")
        return Path(f"/mnt/{drive}/{rest}")
    return Path(s)

def _load_rows_from_excel(xlsx_path: Path) -> List[Dict[str, Any]]:
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Excel file not found: {xlsx_path}")

    wb = load_workbook(str(xlsx_path), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(h or "").strip() for h in rows[0]]
    data_rows = rows[1:]

    # Identify canonical section columns from header, falling back to SECTIONS keys
    canonical_list = list(SECTIONS.keys())
    cols = {name: idx for idx, name in enumerate(header)}

    required_cols = ["pdf_path", "contact_info"]
    for r in required_cols:
        if r not in cols:
            raise ValueError(f"Missing required column '{r}' in Excel header: {header}")

    # Any remaining columns treat as section columns; prefer known canonical order if present
    section_cols_in_sheet = [h for h in header if h not in required_cols]
    # Keep order by canonical_list
    ordered_sections = [s for s in canonical_list if s in section_cols_in_sheet]
    # Add any extra unknown columns at the end
    ordered_sections += [h for h in section_cols_in_sheet if h not in ordered_sections]

    items: List[Dict[str, Any]] = []
    for i, r in enumerate(data_rows):
        if r is None:
            continue
        pdf_path = (r[cols["pdf_path"]] or "").strip() if len(r) > cols["pdf_path"] else ""
        if not pdf_path:
            continue
        contact_raw = (r[cols["contact_info"]] or "") if len(r) > cols["contact_info"] else ""
        contact: Dict[str, Any] = {}
        if isinstance(contact_raw, str) and contact_raw.strip():
            try:
                contact = json.loads(contact_raw)
            except Exception:
                # keep raw string if not JSON
                contact = {"raw": contact_raw}
        elif isinstance(contact_raw, dict):
            contact = contact_raw

        sections: Dict[str, Any] = {}
        for sec in ordered_sections:
            idx = cols.get(sec)
            if idx is None or idx >= len(r):
                continue
            val = r[idx] or ""
            # Section lines were joined with '\n' in batch writer; split to list for UI
            if isinstance(val, str):
                lines = [ln for ln in (val.splitlines()) if ln.strip()]
            else:
                lines = [str(val)] if val else []
            sections[sec] = lines

        items.append({
            "index": i,  # row index (relative to data rows)
            "pdf_path": pdf_path,
            "contact": contact,
            "sections": sections,
        })
    return items

def _resolve_excel_path() -> Path:
    # Allow override via env
    if os.getenv("BATCH_XLSX"):
        p = Path(os.getenv("BATCH_XLSX"))
        if not p.is_absolute():
            p = (ROOT_DIR / p).resolve()
        if p.exists():
            return p

    candidates = [
        ROOT_DIR / DEFAULT_XLSX,
        ROOT_DIR / ALSO_TRY_XLSX,
        ROOT_DIR / FALLBACK_XLSX,
        ROOT_DIR / "outputs" / DEFAULT_XLSX,
    ]
    for p in candidates:
        if p.exists():
            return p

    # last resort: look in root for any xlsx containing 'batch' in name
    for cand in ROOT_DIR.glob("*.xlsx"):
        if "batch" in cand.name.lower():
            return cand
    raise FileNotFoundError(f"Could not find {DEFAULT_XLSX} or {ALSO_TRY_XLSX} or {FALLBACK_XLSX} in {ROOT_DIR}")

def _build_allowed_pdf_map(items: List[Dict[str, Any]]) -> Dict[int, Path]:
    """
    Map index -> absolute Path. Only serve files that appear in the Excel.
    Also translate Windows-style absolute paths to WSL if needed.
    """
    idx_to_path: Dict[int, Path] = {}
    for i, item in enumerate(items):
        raw = item.get("pdf_path", "")
        p = Path(raw)

        # Prefer absolute resolution
        if not p.is_absolute():
            p = (ROOT_DIR / p).resolve()

        # If still missing, try Windows->WSL translation
        if not p.exists():
            alt = _windows_to_wsl_path(raw)
            if alt.is_absolute():
                p = alt.resolve()

        idx_to_path[i] = p
    return idx_to_path

def create_app() -> Flask:
    app = Flask(__name__)

    # Load data once at startup
    xlsx = _resolve_excel_path()
    items = _load_rows_from_excel(xlsx)
    path_map = _build_allowed_pdf_map(items)

    @app.get("/")
    def index() -> Response:
        if not INDEX_HTML.exists():
            return Response("index.html not found at project root.", status=404)
        return send_file(str(INDEX_HTML), mimetype="text/html; charset=utf-8")

    @app.get("/api/rows")
    def api_rows():
        # Return lightweight payload
        payload = []
        for i, item in enumerate(items):
            payload.append({
                "i": i,
                "pdf_path": item["pdf_path"],
                "filename": Path(item["pdf_path"]).name,
                "contact": item.get("contact", {}),
                "sections": item.get("sections", {}),
            })
        return jsonify(payload)

    @app.get("/pdf/<int:i>")
    def serve_pdf_by_index(i: int):
        if i < 0 or i >= len(items):
            abort(404)
        p = path_map.get(i)
        if not p or not p.exists() or not p.suffix.lower() == ".pdf":
            abort(404)
        # Stream PDF
        return send_file(str(p), mimetype="application/pdf", conditional=True)

    return app

if __name__ == "__main__":
    app = create_app()
    # Windows/WSL-friendly localhost run
    app.run(host='0.0.0.0', port=5000, debug=True)

